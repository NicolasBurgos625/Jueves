from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from .forms import TrabajadorForm, PaqueteForm, EditarTrabajadorForm
from .models import Trabajador, Paquete, Notificacion
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.contrib import messages
from openpyxl import Workbook
from openpyxl.styles import Font
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from .models import Paquete
from .forms import PaqueteForm



# Vista de login
def login_view(request):
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            if user.is_superuser:  # Si el usuario es un superusuario
                return redirect('/admin/')  # Redirige al panel de administración
            return redirect('dashboard')  # Si no es superusuario, redirige al dashboard
        else:
            return render(request, 'core/index.html', {'error': 'Credenciales incorrectas'})
    return render(request, 'core/index.html')


# Vista de logout
@login_required
def logout_view(request):
    logout(request)
    return redirect('index')


# Dashboard principal
@login_required
def dashboard(request):
    if hasattr(request.user, 'trabajador'):
        if request.user.trabajador.rol == 'jefe':
            return redirect('jefe_dashboard')
        elif request.user.trabajador.rol == 'repartidor':
            return redirect('repartidor_dashboard')
        elif request.user.trabajador.rol == 'recursos_humanos':
            return redirect('recursos_humanos_dashboard')
        elif request.user.trabajador.rol == 'recepcionista':
            return redirect('recepcionista_dashboard')
    return redirect('logout')


#jefe dashboard
@login_required
def jefe_dashboard(request, rol=None):
    if request.user.trabajador.rol != 'jefe':
        return redirect('logout')  # Asegura que solo el jefe pueda acceder

    # Filtrar empleados según el rol proporcionado, si hay un rol, o mostrar todos los empleados
    if rol:
        empleados = Trabajador.objects.filter(rol=rol)  # Filtrar por rol
    else:
        empleados = Trabajador.objects.all()  # Mostrar todos los empleados si no se selecciona un rol

    # Se definen los empleados por rol para usarlos en la navegación
    repartidores = Trabajador.objects.filter(rol='repartidor')
    recepcionistas = Trabajador.objects.filter(rol='recepcionista')
    rrhh = Trabajador.objects.filter(rol='recursos_humanos')
    
    # Obtener todos los paquetes con los datos necesarios
    paquetes = Paquete.objects.select_related('repartidor').all()  # Asumiendo que el paquete tiene un campo 'repartidor'

    # Contar los paquetes por categorías
    total_paquetes = Paquete.objects.count()  # Total de paquetes
    paquetes_asignados = Paquete.objects.filter(repartidor__isnull=False).count()  # Paquetes asignados
    paquetes_entregados = Paquete.objects.filter(estado='Entregado').count()  # Paquetes entregados
    paquetes_sin_asignar = Paquete.objects.filter(repartidor__isnull=True).count()  # Paquetes sin asignar

    # Pasar los datos al contexto
    context = {
        'empleados': empleados,
        'repartidores': repartidores,
        'recepcionistas': recepcionistas,
        'rrhh': rrhh,
        'paquetes': paquetes,
        'rol': rol,  # Pasamos el rol para saber cuál está seleccionado
        'total_paquetes': total_paquetes,
        'paquetes_asignados': paquetes_asignados,
        'paquetes_entregados': paquetes_entregados,
        'paquetes_sin_asignar': paquetes_sin_asignar,
    }
    return render(request, 'core/jefe.html', context)

# Descargar tabla en Word para el jefe
@login_required
def descargar_tabla(request):
    if request.user.trabajador.rol != 'jefe':
        return redirect('dashboard')

    # Crear el objeto Workbook (libro de trabajo)
    wb = Workbook()
    ws = wb.active
    ws.title = "Paquetes"

    # Definir los encabezados de la tabla
    headers = ["Código", "Destinatario", "Dirección", "Estado", "Repartidor"]

    # Establecer el estilo de fuente para los encabezados
    bold_font = Font(bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = bold_font

    # Filtrar los paquetes que tienen un repartidor asignado
    paquetes = Paquete.objects.filter(repartidor__isnull=False)

    # Llenar la tabla con los datos de los paquetes
    for row_num, paquete in enumerate(paquetes, start=2):
        ws.cell(row=row_num, column=1, value=paquete.codigo)
        ws.cell(row=row_num, column=2, value=paquete.destinatario)
        ws.cell(row=row_num, column=3, value=paquete.direccion)
        ws.cell(row=row_num, column=4, value=paquete.estado)
        # Si el repartidor está asignado, se añade su nombre
        if paquete.repartidor:
            ws.cell(row=row_num, column=5, value=f"{paquete.repartidor.nombre} {paquete.repartidor.apellido}")
        else:
            ws.cell(row=row_num, column=5, value="Sin asignar")

    # Preparar la respuesta para descargar el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=tabla_paquetes.xlsx'

    # Guardar el archivo Excel en la respuesta
    wb.save(response)
    return response


# Dashboard de recursos humanos
@login_required
def recursos_humanos_dashboard(request):
    trabajadores = Trabajador.objects.all()
    return render(request, 'core/recursos_humanos.html', {'trabajadores': trabajadores})


@login_required
def repartidor_dashboard(request):
    # Obtener los paquetes del repartidor
    paquetes = Paquete.objects.filter(repartidor=request.user.trabajador)

    # Contar las notificaciones activas
    notificaciones_activas = request.user.trabajador.notificaciones.filter(activo=True)

    # Pasar los datos al template
    return render(request, 'core/repartidor.html', {
        'paquetes': paquetes,
        'notificaciones_activas': notificaciones_activas,  # Pasamos las notificaciones activas
        'contador_notificaciones': notificaciones_activas.count(),  # Contador de notificaciones
    })



@login_required
def recepcionista_dashboard(request):
    # Obtener todos los repartidores
    repartidores = Trabajador.objects.filter(rol='repartidor')

    # Obtener todos los paquetes, que incluyen tanto los asignados como los no asignados
    paquetes = Paquete.objects.all()

    # Obtener todos los trabajadores
    trabajadores = Trabajador.objects.all()

    # Obtener el filtro para mostrar las diferentes opciones de tablas
    filtro_paquetes = request.GET.get('filtro', 'asignados')  # Por defecto se muestra 'asignados'

    # Validar que el filtro es uno de los valores válidos
    if filtro_paquetes not in ['asignados', 'sin_asignar']:
        filtro_paquetes = 'asignados'

    if filtro_paquetes == 'sin_asignar':
        # Filtrar solo los paquetes sin asignar
        paquetes = paquetes.filter(repartidor__isnull=True)
    elif filtro_paquetes == 'asignados':
        # Filtrar solo los paquetes asignados
        paquetes = paquetes.filter(repartidor__isnull=False)

    # Contar los paquetes por categorías
    total_paquetes = Paquete.objects.count()  # Total de paquetes
    paquetes_asignados = Paquete.objects.filter(repartidor__isnull=False).count()  # Paquetes asignados
    paquetes_entregados = Paquete.objects.filter(estado='Entregado').count()  # Paquetes entregados
    paquetes_sin_asignar = Paquete.objects.filter(repartidor__isnull=True).count()  # Paquetes sin asignar

    # Pasar los datos al contexto
    context = {
        'repartidores': repartidores,
        'paquetes': paquetes,
        'trabajadores': trabajadores,
        'filtro': filtro_paquetes,  # Pasamos el filtro elegido
        'total_paquetes': total_paquetes,
        'paquetes_asignados': paquetes_asignados,
        'paquetes_entregados': paquetes_entregados,
        'paquetes_sin_asignar': paquetes_sin_asignar,
    }

    return render(request, 'core/recepcionista.html', context)


# Editar datos personales del trabajador
@login_required
def editar_trabajador(request):
    trabajador = request.user.trabajador
    if request.method == "POST":
        form = EditarTrabajadorForm(request.POST, instance=trabajador)
        if form.is_valid():
            trabajador = form.save(commit=False)
            trabajador.save()  # Guardar los cambios
            return redirect('dashboard')  # Redirigir al dashboard
    else:
        form = EditarTrabajadorForm(instance=trabajador)
    
    return render(request, 'core/editar_trabajador.html', {'form': form})

# Vista para ver los detalles de un paquete
@login_required
def paquete_detail(request, paquete_id):
    paquete = get_object_or_404(Paquete, id=paquete_id)
    
    if request.method == 'POST':
        form = PaqueteForm(request.POST, instance=paquete)
        if form.is_valid():
            form.save()
            return redirect('recepcionista_dashboard')  # Redirigir después de guardar los cambios
    else:
        form = PaqueteForm(instance=paquete)
    
    return render(request, 'core/paquete_detail.html', {'paquete': paquete, 'form': form})

@login_required
def editar_paquete(request, paquete_id):
    paquete = get_object_or_404(Paquete, id=paquete_id)

    if request.method == 'POST':
        # Recoger datos del formulario
        paquete.codigo = request.POST.get('codigo')
        paquete.destinatario = request.POST.get('destinatario')
        paquete.direccion = request.POST.get('direccion')
        paquete.estado = request.POST.get('estado')
        paquete.save()

        return redirect('paquete_detail', paquete_id=paquete.id)  # Redirigir a los detalles del paquete

    return render(request, 'core/paquete_edit.html', {'paquete': paquete})

@login_required
def asignar_paquete(request, paquete_id):
    paquete = get_object_or_404(Paquete, id=paquete_id)

    if request.method == 'POST':
        # Obtener el ID del repartidor seleccionado
        repartidor_id = request.POST.get('repartidor_id')
        if repartidor_id:
            # Buscar al repartidor por ID
            repartidor = get_object_or_404(Trabajador, id=repartidor_id)

            # Asignar el paquete al repartidor
            paquete.repartidor = repartidor
            paquete.estado = 'Asignado'  # Actualizamos el estado del paquete a 'Asignado'
            paquete.save()

            # Crear la notificación para el repartidor
            mensaje = f"Se te ha asignado el paquete {paquete.codigo}."
            notificacion = Notificacion(trabajador=repartidor, mensaje=mensaje)
            notificacion.save()

            return redirect('recepcionista_dashboard')  # Redirigir al dashboard de recepcionista

    return redirect('recepcionista_dashboard')  # Redirigir en caso de error o si no es un POST





@login_required
def registrar_trabajador(request):
    if request.method == 'POST':
        form = TrabajadorForm(request.POST)
        if form.is_valid():
            try:
                trabajador = form.save()
                messages.success(request, f"Trabajador {trabajador.nombre} {trabajador.apellido} registrado exitosamente.")
                return redirect('rrhh_dashboard')
            except Exception as e:
                messages.error(request, f"Error al registrar trabajador: {str(e)}")
    else:
        form = TrabajadorForm()
    
    return render(request, 'core/registrar_trabajador.html', {'form': form})

# Vista para actualizar el estado del paquete a "Entregado"
@login_required
def actualizar_estado_paquete(request, paquete_id):
    # Buscar el paquete por ID
    paquete = get_object_or_404(Paquete, id=paquete_id)

    # Verificar si el usuario es un repartidor y si el paquete está asignado a él
    if request.user.trabajador.rol == 'repartidor' and paquete.repartidor == request.user.trabajador:
        # Actualizar el estado del paquete a 'Entregado'
        paquete.estado = 'Entregado'
        paquete.save()
        messages.success(request, f"Paquete {paquete.codigo} marcado como entregado.")
    else:
        messages.error(request, "No tienes permisos para actualizar este paquete o el paquete no te está asignado.")
    
    # Redirigir al dashboard del repartidor
    return redirect('repartidor_dashboard')

@login_required
def eliminar_trabajador(request, trabajador_id):
    if request.user.trabajador.rol != 'recursos_humanos':
        return redirect('dashboard')  # Asegúrate de que solo los usuarios de RH puedan eliminar trabajadores

    trabajador = get_object_or_404(Trabajador, id=trabajador_id)
    trabajador.delete()
    return redirect('recursos_humanos_dashboard')  # Redirigir a la página de recursos humanos


@login_required
def modificar_trabajador(request, trabajador_id):
    # Obtener al trabajador específico usando su ID
    trabajador = get_object_or_404(Trabajador, id=trabajador_id)

    if request.method == "POST":
        # Si es una solicitud POST, procesamos el formulario
        form = TrabajadorForm(request.POST, instance=trabajador)
        if form.is_valid():
            form.save()  # Guardamos los cambios
            return redirect('recursos_humanos_dashboard')  # Redirigimos al dashboard de recursos humanos
    else:
        # Si es GET, cargamos el formulario con los datos actuales del trabajador
        form = TrabajadorForm(instance=trabajador)

    # Renderizamos la plantilla y le pasamos el formulario
    return render(request, 'core/modificar_trabajador.html', {'form': form, 'trabajador': trabajador})


@csrf_exempt  # Asegúrate de tener esto si usas AJAX
def desactivar_notificacion(request, notificacion_id):
    if request.method == 'POST':
        try:
            notificacion = request.user.trabajador.notificaciones.get(id=notificacion_id)
            notificacion.activo = False  # Desactivamos la notificación
            notificacion.save()
            return JsonResponse({'success': True})
        except Notificacion.DoesNotExist:
            return JsonResponse({'success': False})
    return JsonResponse({'success': False})


def obtener_datos_actualizados(request):
    # Obtener los paquetes y las notificaciones
    paquetes = Paquete.objects.filter(repartidor=request.user.trabajador, estado='Pendiente')
    notificaciones_activas = Notificacion.objects.filter(trabajador=request.user.trabajador, activa=True)

    # Preparamos los datos para enviar como JSON
    paquetes_data = [{
        'codigo': paquete.codigo,
        'destinatario': paquete.destinatario,
        'direccion': paquete.direccion,
        'estado': paquete.estado,
    } for paquete in paquetes]

    notificaciones_data = [{
        'mensaje': notificacion.mensaje,
        'fecha': notificacion.fecha.strftime('%Y-%m-%d %H:%M:%S'),
    } for notificacion in notificaciones_activas]

    # Devolver los datos en formato JSON
    return JsonResponse({
        'paquetes': paquetes_data,
        'notificaciones': notificaciones_data,
    })





































